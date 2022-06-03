import logging
logging.basicConfig(level=logging.INFO,
#filename= 'output.txt',
format='%(asctime)s - %(levelname)s - %(message)s')
logging.debug("Start of program")
#logging.disable(logging.CRITICAL)  #end all logging messages
def theFirst():
    complex_number = 3 + 2j

    #execute a python file from code
    #execfile("prelimTractor.py")

    #evaluate a string form of code [check the functions for variables input]
    eval("s = 15")

    coolString = ("The letter of the day is %c. My favourite word is %s. My favourite number is %d. A random number is %3.2f."
                  % ('D', 'Mango', 7, 1024.629))

def writeTxt(info="Nothing", nm='text'):
    textFile = open('%s.txt' % nm, 'wb')  # 'wb' - Write file
    print(textFile.mode)
    print(textFile.name)
    byteData = bytes(info, 'UTF-8')
    lenChar = textFile.write(byteData)
    textFile.close()

def extractTxt(nm='text'):
    name = '%s.txt' % nm
    textFile = open(name, 'r+')
    text = textFile.read()
    import os
    os.remove(name)
    print("The text, '%s' has been extracted" % name)
    return text

def user_input_number():
    import sys
    target_int = input("How many intergers?")

    try:
        target_int = int(target_int)
    except ValueError:
        sys.exit("The number entered was not an interger. Enter an interger")

    ints = list()
    count = 0

    while count < target_int:
        new_int = input("Please enter interger {0}:".format(count + 1))
        isint = False
        try:
            new_int = int(new_int)
        except:
            print("You must enter an interger")

        if isint:
            ints.append(new_int)
            count += 1

    print("Using a for loop")
    for number in ints:
        print(str(number))

    print("Using a while loop")
    total = len(ints)
    count = 0
    while count > total:
        print(str(ints[count]))
        count += 1

# Ask the user to input 2 values and store them in variables; num1, num2


def questions():
    num1, num2 = input('Enter 2 numbers: ').split()

    # Convert strings into regular numbers, Interger
    num1 = float(num1)
    num2 = float(num2)

    # Add the values entered and store in, sum
    sum = num1 + num2

    # Subtract values and store in different
    different = num1 - num2

    # Multiply the values and store in product
    product = num1 * num2

    # Divide the values and store in quotient
    quotient = num1 / num2

    # Use modulus on the values to find the remainder
    remainder = num1 % num2

    # Print the results
    print(f'{num1} + {num2} = {sum}')
    print('{} - {} = {}'.format(num1, num2, different))
    print(f'{num1} * {num2} = {product}')
    print(f'{num1} / {num2} = {quotient}')
    print(f'{num1} % {num2} = {remainder}')

def loopin():
    from random import randrange
    #initiation
    rand_num = randrange(1, 51)
    i = 1
    while (i != rand_num):
        i += 1
    print(f'The random value is : {rand_num}')

def pineTree():
    while True:
        try:
            rows = int(input("How tall is the tree?"))
            break
        except ValueError:
            print("You didn't enter a number")
        except:
            print("An unknown error occured")
    rows = int(rows)
    stump = 1

    if rows <= 0:
        print("The wrong value was entered or the tree is too short.")
        pineTree()
        return None
    elif rows == stump:
        print("The tree is too short.")
        pineTree()
        return None
    elif rows > 1:
        print(f"Here is the {rows} rows-tall tree.")
        i = rows
        while i > stump:
            print(" " * i, "#" * (1 + (rows - i)*2))
            i -= 1
        while i > 0:
            print(" " * rows, "#")
            i -= 1

def guessNumber():
    import random
    # request user input
    while True:
        # generate a random number from 1 to 10
        sol = random.randrange(1, 11)
        try:
            ans = input("Please guess a number from 1 to 10. ")
        except ValueError:
            print("You entered a wrong value ")
            continue

        if ans is sol:
            print(f"You are Correct!!! The answer is {sol}")
            break
        else:
            print(
                f"Sorry, that was incorrect. The answer is {sol}. Guess again")

def bubbleSort(lowerLimit=1, upperLimit=10):
    #not working
    import random
    while True:
        try:
            quantity = int(
                input("How many random numbers do you want to sort? "))
            break
            #ranger = input("Enter the range for the random numbers").split()
        except ValueError:
            print("You entered a wrong value. Try Again. ")
    numbers = list(range(quantity))
    # TBD: error check for lowerLimit and upperLimit
    # create random list
    for i in range(quantity):
        numbers[i] = random.randrange(lowerLimit, upperLimit + 1)
    print("Scrambled: ", numbers)
    # TIME TO SORT
    i = quantity
    swapped = True
    while(swapped):
        for d in range(1, i):
            before = numbers[d - 1]
            after = numbers[d]
            if before > after:
                numbers[d - 1] = after
                numbers[d] = before
                swapped = True
        i -= 1
        if(i < 2):
            swapped = False
    print("Unscarambled: ", numbers)
    return tuple(numbers)

def multiplicationTable(size=12):
    #table = [list(range(size)) for i in range(size)]
    table = [[0] * size for i in range(size)]   #list comprehension
    '''
        The '[0] * size' returns a list of zeros the size of 'size'.
        This is then given to each iteration in the range.
    '''
    toPrint = ""
    for a in range(size):
        for b in range(size):
            num = (a+1) * (b+1)
            table[a][b] = num
            #toPrint += str(num) + ", "
            strn = str(num)
            toPrint += strn + "," + (' ' * (4 - len(strn)))
        toPrint += "\n"
    # Remove the last comma
    toPrint = toPrint[:-1]
    print(toPrint)
    return table

def customerLister():

    customerNames = []

    while True:
        ask = input("Enter Customer (Yes/No) : ")

        if 'y' not in ask.lower():
            break

        names = input("Enter Customer Name: ")
        # TBD: Error capture of name splitting
        firstN, lastN = names.split()

        customerNames.append({'first': firstN, 'last': lastN})

    print("These are the list of customers.")
    for cust in customerNames:
        print(f"{cust['first']} {cust['last']}")


def writeNstuff(choice = 0):
    import os

    def cleanName(strung):
        pass
        #clean-up the string here to be used in windows name

    choice = choice if choice is not 0 else input("""Enter the number of the choosen function. 
    1. Write text to text file. 
    2. Read text in text file. 
    3. Change name of text file.
    4. 
    """)

    #myFile.readable()
    #myFile.readline()
    #os.getcwd()

    #Filter choice for errors
    try:
        choice = int(choice)
    except:
        print("There was a problem with the value entered.")

    #choosen function
    if choice == 1:
        strung = input("What is the information you want to save? \n")
        name = input("What is the name of the file you want to save to? \n")
        #TBD: Clean-up 'name' here

        with open(name + ".txt", mode="w", encoding ="utf-8") as myFile:
            myFile.write(strung)

    elif choice == 2:
        name = input("What is the name of the file? \n")
        
        with open(name + ".txt", "r", "utf-8") as myFile:
            text = myFile.read()
            print(text)
    
    elif choice == 3:
        oldName = input("What is the name of the file? \n")
        newName = input("What do you want to name the file? \n")

        os.rename(oldName, newName)

    
def peopleHere():
    from secondy import person

    people1 = person("Darnell Baird", 181, 80)

def chrTest(limit = 100, textName= 'idgaf', printQ = False):

    for i in range(limit):
        str = chr(i)
        if len(str) > 1:
            string = f"The chr() of {str} is {i} which is a string. Its length is {len(str)}."

        elif len(str) == 1:
            string = f"The chr() of {str} is {i} and is a single character."

        else:
            string = f"Something is wrong with {str}"

        with open(textName + ".txt", mode = "a+", encoding = "utf-8") as myFile:
            myFile.write(string + "\n")

    print(string) if printQ else None

def modFactorial(num):
    import math
    i = 1
    while num > math.factorial(i):
        i += 1
    i -= 1
    rem = num - math.factorial(i)
    print(f"The factorial is {i} and the remainder is {rem}")
    return (i, rem)

def enumTest():
    list = ["sheep", "goat", "cow", "sheep"]
    enu = enumerate(list, 3)
    for c, v in enu:
        print("index", c, "value", v)
    #print("The type", type(enu[0]))
    print(dir(enu))

def tqdmTest():
    from time import sleep
    from tqdm import tqdm

    progress_bar = tqdm(list(string.ascii_lowercase))
    for letter in progress_bar:
        progress_bar.set_description(f'Processing {letter}...')
        sleep(0.09)

def retirement(years = 40, percentInterestRate = 5, deposit = 5000):
    choice = 0
    balance = 0
    months = int(years * 12)
    interestRate = percentInterestRate/100
    
    if choice == 0:
        for monthCount in range(months):
            balance += deposit
            if monthCount % 12 == 11:
                balance *= (1 + interestRate)
    elif choice == 1:
        import numpy as np
        balance = np.fv(interestRate/12, months, -deposit, balance)
    
    financialStatement = f"The balance is {balance}. This took {months} months at an interest rate of {percentInterestRate}%"
    print(financialStatement)

def directoryPlayer():
    import pathlib
    #Initiation
    currentDir = pathlib.Path()
    path = pathlib.Path("Books")
    path2 = pathlib.Path("Goat Skin")
    
    exist = path.exists()
    path2.mkdir()
    path2.rmdir()
    #Search for files in directory
    for file in currentDir.glob("*.txt"):
        print(file)

def excelPlay(which = 1):
    import openpyxl
    wb = openpyxl.load_workbook("logging times.xlsx")
    sheet = wb["Sheet1"]
    cell = sheet['a1']
    cellAlt = sheet.cell(1,1)

    if which == 1:
        hrs = 0
        maxRow = sheet.max_row
        for x in range(1, maxRow + 1):
            cell = sheet.cell(x, 3) #row, column
            val = cell.value
            #print(cell.value, type(cell.value))
            if type(val) == type(2):
                hrs += val
        print(hrs, "hours")

def genCountdown(startFrom = 10):
    print("The countdown begins from ", startFrom)
    print("Countdown initiated.")
    while startFrom > 0:
        print(startFrom)
        yield startFrom
        startFrom -= 1

    #end save
    #wb.save("revised" + wb.name)

class A(object):
    def __init__(self, a, b, c, d, e, f):
        self.__dict__.update({k: v for k, v in locals().items() if k != 'self'})

def genTest(size=3):
    gen = (((i + 1) + (size*abb) for i in range(size)) for abb in range(size))
    for subGen in  gen:
        for subVal in subGen:
            print(subVal)

def iterateThroughList(listy):
    for i in listy:
        yield i

def jsonTest():
    import json
    choice = 0
    ans = None
    if choice == 0:
        #to json
        x = {
            "name": "Darnell",
            "age":30,
            "city":"Port-of-Spain",
            "car":[
                {"model":"Laurel Medalist", "mpg": 20.3},
                {"model":"Corolla", "mpg": 23}
            ]
        }
        print(json.dumps(x))
    elif choice == 1:
        #json string format to dictionary
        str = '{ "name":"John", "age":30, "city":"Port of Spain"}'
        ans = json.loads(str)
    else:
        print("I 'ent doing nunnin'")

def playKivy():
    import re, sys, copy, time
    import numpy as np
    import kivy
    from functools import partial
    from kivy.app import App
    from kivy.uix.button import Button
    from kivy.uix.togglebutton import ToggleButton
    from kivy.uix.label import Label
    from kivy.uix.widget import Widget
    from kivy.uix.textinput import TextInput
    from kivy.uix.relativelayout import RelativeLayout
    from kivy.uix.gridlayout import GridLayout
    from kivy.uix.scrollview import ScrollView
    from kivy.uix.screenmanager import ScreenManager, Screen
    from kivy.core.window import Window
    from kivy.clock import Clock

    class Dong(App):
        def __init__(self, **kwargs):
            super().__init__(**kwargs)

        def going(self, instance, value):
            #not good
            a = 5
            while a > 0:
                time.sleep(1)
                self.info = str(a) + " " + str(value)
                a -= 1

        def kivyIntervala(self, instance, value):
            Clock.schedule_interval(self.my_callback, 0.5)

        def my_callback(self, dt):
            self.info.text = "Dog"
            if dt > 5:
                return False
            
        def build(self, **kwargs):
            #widgets
            sm = ScreenManager()
            rid = GridLayout(cols=2)
            widge = Label(text="Hi!")
            butt = Button(text="Click Me!", size_hint=(0.2,0.5))
            butt2 = Button(text="Organic!")
            self.info = Label(text="Nunnin'")

            #addwidget
            rid.add_widget(widge)
            rid.add_widget(butt)
            rid.add_widget(self.info)

            #callback
            #butt.bind(state=self.going)
            butt2.bind(state=self.kivyIntervala)

            #return widget
            return rid

    obj = Dong()
    obj.run()


def testVarInFuncFromFunc():
    '''
    What happens to the variable in a function if I return the function that reference that variable?
    '''
    def outFunc():
        permVar1 = 1
        def returningFunc(num):
            permVar = permVar1
            print("The number is", num, ". The permanent value is", permVar, "and and I'll update it now.")
            permVar = num
        return returningFunc
    #run the code
    rf = outFunc()  #inferred
    rf(5)
    rf(3)

def simultaneousThreads():
    import time, threading
    from pynput.keyboard import Key, Listener
    from pynput.mouse import Button as mouseButton, Controller as mouseController
    class Hold:
        loop = True
        character = ''
        count = 0
        def on_press(self, key):
            print('{0} pressed'.format(key))
            self.character = key
        def on_release(self, key):
            print('{0} release'.format(key))
            if key == Key.esc:
                # Stop listener
                self.loop = False
                return False
        def one(self):
            print("One started.")
            # Collect events until released
            with Listener(on_press=self.on_press,on_release=self.on_release) as listener:
                listener.join()
        def two(self):
            print("Two started.")
            while self.loop:
                time.sleep(1)
                self.count += 1
                print(f"This is {self.character} @ {self.count}.")
            print("This is the end of Two.")
        def run(self):
            thread1 = threading.Thread(target=self.one)
            thread2 = threading.Thread(target=self.two)
            thread1.start()
            thread2.start()
    obj = Hold()
    obj.run()
def concurrentProcesses():
    #run several processes and then get the return values of each when completed
    import time, multiprocessing, concurrent
    def doSomething(waitTime=1):
        print(f'Sleeping {waitTime} seconds.')
        time.sleep(waitTime)
        return f'Done Sleeping {waitTime} seconds.'
    with concurrent.futures.ProcessPoolExecutor() as executor:
        results = [executor.submit(doSomething, 1) for i in range(8)]
        #get actual results now
        for f  in concurrent.futures.as_completed(results):
            print(f.result())

def decoratorCheckTime(func):
    import time
    def wrapper(*arg, **kwargs):
        startTime = time.monotonic()
        func(*arg, **kwargs)
        logging.info(f"This is how long it took. {time.monotonic() - startTime}")
    return wrapper

def readURLOffWebpage():
    from bs4 import BeautifulSoup
    from urllib.request import urlopen
    import re, requests

    html_page = urlopen("https://tractortestlab.unl.edu/john-deere")
    soup = BeautifulSoup(html_page)
    for link in soup.findAll('a', attrs={'href': re.compile("^http://")}):
        print(link.get('href'))
    
    r = requests.get("https://tractortestlab.unl.edu/john-deere")
    str = r.text
    with open("blax" + ".txt", mode="w", encoding ="utf-8") as myFile:
            myFile.write(str)

def mySearchEngine():
    import socket
    mySocket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    mySocket.connect(("data.pr4e.org",80))
    cmd = "GET http://data.pr4e.org/romeo.txt HTTP/1.0\n\n".encode()    #from string format to byte
    mySocket.send(cmd)

    while True:
        data = mySocket.recv(512)   #receive 512 characters at a time.
        if (len(data) < 1):
            break
        print(data.decode())    #from byte format to string
    mySocket.close()

def readURLOffWebpage3():
    import requests, re
    from bs4 import BeautifulSoup
    url = "https://www.google.com/"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    #pattern = r"(https:\/\/)(www\.)?[\da-z\.-]+\.([a-z\.]{2,6})"
    pattern = r"https\/\/www\.[a-z]+\.com"
    liss = re.findall(pattern, str(soup))
    [print(x) for x in liss]
    return
    for url in soup.li.find_all("a"):
        print(url.get('href'))
    for base in soup.find_all("base"):
        print(base.get('href'))

@decoratorCheckTime
def readURLOffWebpage2(motherPageURL="https://tractortestlab.unl.edu"):
    import requests, re, PyPDF2
    from multiprocessing import Process, Lock
    from bs4 import BeautifulSoup
    logging.level = logging.INFO

    nameReplace = lambda str: re.sub(r"\W", " ", str)
    def soupUp(locki, docURL):
        try:
            subResponse = requests.get(motherPageURL+docURL)
            subSoup = BeautifulSoup(subResponse.content, 'html.parser')
            for subURL in subSoup.find_all('a'):
                pdfURL = subURL.get('href')
                if 'pdf' in pdfURL:
                    #print(pdfURL)
                    pdfResponse = requests.get(motherPageURL+pdfURL)
                    pdfName = nameReplace(pdfURL) + ".pdf"
                    with open(pdfName, "wb") as myData:
                        myData.write(pdfResponse.content)
                    openPDFFile = open(pdfName,'rb')
                    readPDF = PyPDF2.PdfFileReader(openPDFFile)
                    if readPDF.isEncrypted:
                        readPDF.decrypt("")
                    try:
                        print(pdfName)
                        locki.acquire()     #lock down occurs here. For sequencing purposes
                        print(readPDF.getPage(0).extractText())
                        print("***\n"*3)
                    finally:
                        locki.release()

        except requests.exceptions.RequestException as e:
            logging.debug(e)

    lock = Lock()
    response = requests.get(motherPageURL)
    soup = BeautifulSoup(response.content, 'html.parser')
    #iterate through each href url on mother page
    for url in soup.find_all("a"):
        try:
            link = url.get('href')
            Process(target=soupUp,args=(lock, link)).start()
        except Exception as h:
            logging.debug("Missed it. " + str(h))
        

@decoratorCheckTime
def pdfOnlineTest():
    import requests, re, PyPDF2, json
    from bs4 import BeautifulSoup
    motherPage = "https://tractortestlab.unl.edu/documents/T6020-Elite.pdf"
    response = requests.get(motherPage)

    with open("my_pdf.pdf", 'wb') as my_data:
        my_data.write(response.content)
    open_pdf_file = open("my_pdf.pdf", 'rb')
    read_pdf = PyPDF2.PdfFileReader(open_pdf_file)
    if read_pdf.isEncrypted:
        read_pdf.decrypt("")
    str = ""
    #print(read_pdf.getPage().extractText())
    dic = read_pdf.getDocumentInfo()   #pdf info
    #str = json.dumps(dic)
    #print(str)
    #[print(read_pdf.getPage(x).extractText()) for x in range(read_pdf.getNumPages())]
    #print(read_pdf.extractText())
    print(dir(read_pdf))

def pdfminerTest():
    from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
    from pdfminer.converter import TextConverter
    from pdfminer.layout import LAParams
    from pdfminer.pdfpage import PDFPage
    from io import BytesIO
    import os

    path = os.path.join(os.getcwd(), 'my_pdf.pdf')
    manager = PDFResourceManager()
    retstr = BytesIO()
    layout = LAParams()
    device = TextConverter(manager, retstr, laparams=layout)
    filepath = open(path, 'rb')
    interpreter = PDFPageInterpreter(manager, device)

    for page in PDFPage.get_pages(filepath, check_extractable=True):
        #interpreter.process_page(page)
        print(page)

    text = retstr.getvalue()

    filepath.close()
    device.close()
    retstr.close()
    return text

if __name__ == '__main__':
    #playKivy()
    #readURLOffWebpage2()
    #pdfOnlineTest()
    pdfminerTest()





        
